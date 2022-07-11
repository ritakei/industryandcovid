import subprocess as sp
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
import os


    
def printer(data) -> None:#データ繰り返し表示
    for row in data:
        for cell in row:
            cell_value = cell.value
            if cell_value is not None:
                print(cell.coordinate, cell_value)

def listin(data) -> None:#リスト
    l = []
    for row in data:
        for cell in row:
            cell_value = cell.value
            if cell_value is not None:
                l.append(cell_value)
    return l

def main():   

    if not os.path.exists('park_sales.xlsx'):
        sp.call("curl -tlsv1.2 -v -O -L https://github.com/ritakei/industryandcovid/raw/main/park_sales.xlsx " ,shell=True)
        park_wb = openpyxl.load_workbook("park_sales.xlsx")
        
    if not os.path.exists('internet_service_sales.xlsx'):
        sp.call("curl -tlsv1.2 -v -O -L https://github.com/ritakei/industryandcovid/raw/main/internet_service_sales.xlsx " ,shell=True)
        net_wb = openpyxl.load_workbook("internet_service_sales.xlsx")
            
    if not os.path.exists('newly_confirmed_cases_daily.xlsx'):
        sp.call("curl -tlsv1.2 -v -O -L https://github.com/ritakei/industryandcovid/raw/main/newly_confirmed_cases_daily.xlsx " ,shell=True)
        daily_cases_wb = openpyxl.load_workbook("newly_confirmed_cases_daily.xlsx")

    #openpyxl           
    # Excel読み込み
    park_wb = openpyxl.load_workbook("park_sales.xlsx")
    net_wb = openpyxl.load_workbook("internet_service_sales.xlsx")
    daily_cases_wb = openpyxl.load_workbook("newly_confirmed_cases_daily.xlsx")

    # シートの指定
    park_st = park_wb["month"]
    net_st = net_wb["month"]
    daily_cases_st = daily_cases_wb["newly_confirmed_cases_daily"]

    # データ取得
  
    park_month_range = park_st["C255":"C282"]#月表示 
    park_range = park_st["E255":"E282"]#売上

    net_month_range = net_st["C158":"C185"]#月表示 
    net_range = net_st["E158":"E185"]#売上

    cases_date_range = daily_cases_st["AY2":"AY29"]#日表示
    daily_cases_range = daily_cases_st["BA2":"BA29"]#感染者数

    park_month = listin(park_month_range)
    park_num = listin(park_range)

    net_month = listin(net_month_range)
    net_num = listin(net_range)

    cov_date = listin(cases_date_range)
    cov_num = listin(daily_cases_range)

    #pandas
    d_park = {'park_month': park_month, 'park_num': park_num}
    df_park = pd.DataFrame(data=d_park)
    df_park["park_month"] = pd.to_datetime(df_park["park_month"], format="%Y年%m月").dt.strftime("%Y-%m")

    d_net = {'net_month': net_month, 'net_num': net_num}
    df_net = pd.DataFrame(data=d_net)
    df_net["net_month"] = pd.to_datetime(df_net["net_month"], format="%Y年%m月").dt.strftime("%Y-%m")

    d_cov = {'cov_date': cov_date, 'cov_num': cov_num}
    df_cov = pd.DataFrame(data=d_cov)
    df_cov["cov_date"] = pd.to_datetime(df_cov["cov_date"], format="%Y年%m月").dt.strftime("%Y-%m")

    industry_df = pd.concat([df_park, df_cov,df_net["net_num"]], axis=1)
    print(industry_df)
    
    #matplotlib
    ax_log = plt.subplot(2, 1, 1)
    ax_num = plt.subplot(2, 1, 2)

    ax_log.set_title("Relationship between COVID-19 and Japanese industry(log)")
    
    ax_log.plot(df_park["park_num"],'b.--', label='park')
    ax_log.plot(df_net["net_num"],'y.--', label='internet')
    ax_log.plot(df_cov["cov_num"],'r.-',label='cases')

    ax_log.set_xlabel("Month")
    ax_log.set_xticks([0,1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27], ['20.1','20.2','20.3','20.4','20.5','20.6','20.7','20.8','20.9','20.10','20.11','20.12','21.1','21.2','21.3','21.4','21.5','21.6','21.7','21.8','21.9','21.10','21.11','21.12','22.1','22.2','22.3','22.4'])

    ax_log.set_ylabel("Sales (million)/Number of infected people")
    ax_log.set_yscale('log')
    
    ax_log.grid(True)
    ax_log.legend()

    ax_num.set_title("Relationship between COVID-19 and Japanese industry")

    ax_num.plot(df_park["park_num"],'b.--', label='park')
    ax_num.plot(df_net["net_num"],'y.--', label='internet')
    ax_num.plot(df_cov["cov_num"],'r.-',label='cases')

    ax_num.set_xlabel("Month")
    ax_num.set_xticks([0,1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27], ['20.1','20.2','20.3','20.4','20.5','20.6','20.7','20.8','20.9','20.10','20.11','20.12','21.1','21.2','21.3','21.4','21.5','21.6','21.7','21.8','21.9','21.10','21.11','21.12','22.1','22.2','22.3','22.4'])

    ax_num.set_ylabel("Sales (million)/Number of infected people")
    ax_num.set_yticks([i for i in range(0, 2900000, 150000)])
    ax_num.ticklabel_format(style='plain',axis='y')
    ax_num.grid(True)
    ax_num.legend()
    
    plt.xlabel("Month")
    plt.ylabel("Sales (million)/Number of infected people")
    plt.tight_layout()
    plt.show()#出力

if __name__ == "__main__":
       main()
